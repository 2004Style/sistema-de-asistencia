"""
Reports service for attendance system
Generates reports in PDF and Excel formats
Requerimientos #11-15
"""
from typing import List, Optional, Dict, Any
from datetime import datetime, date, timedelta
from sqlalchemy.orm import Session
from sqlalchemy import and_, func, extract
from pathlib import Path
import io

# Excel generation
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# PDF generation
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT

from src.asistencias.model import Asistencia, EstadoAsistencia
from src.users.model import User
from src.roles.model import Role
from src.horarios.model import Horario
from src.config.settings import get_settings
from src.email.service import email_service
import logging

logger = logging.getLogger(__name__)
settings = get_settings()


class ReportesService:
    """Service for generating attendance reports"""
    
    def __init__(self):
        self.reports_dir = Path(settings.REPORTS_DIR)
        self.reports_dir.mkdir(parents=True, exist_ok=True)
    
    def _preparar_adjuntos(self, archivos: Dict[str, str]) -> List[Dict[str, Any]]:
        """
        Convert files dictionary to attachments list format for email service
        
        Args:
            archivos: {"excel": "path/to/file.xlsx", "pdf": "path/to/file.pdf"}
        
        Returns:
            [{"filename": "report.xlsx", "path": "path/to/file.xlsx"}, ...]
        """
        adjuntos = []
        for tipo, ruta in archivos.items():
            if ruta and Path(ruta).exists():
                adjuntos.append({
                    "filename": Path(ruta).name,
                    "path": ruta
                })
        return adjuntos
    
    # ========================
    # DATA COLLECTION METHODS
    # ========================
    
    def _calcular_datos_asistencia(
        self,
        db: Session,
        user_id: Optional[int],
        fecha_inicio: date,
        fecha_fin: date
    ) -> List[Dict[str, Any]]:
        """
        Calculate attendance data for reports
        """
        try:
            query = db.query(
                Asistencia,
                User.name,
                User.codigo_user,
                User.email
            ).join(User, Asistencia.user_id == User.id)
            
            if user_id:
                query = query.filter(Asistencia.user_id == user_id)
            
            query = query.filter(
                and_(
                    Asistencia.fecha >= fecha_inicio,
                    Asistencia.fecha <= fecha_fin
                )
            ).order_by(Asistencia.fecha.desc(), User.name)
            
            asistencias = query.all()
            
            # Process data
            datos = []
            for asistencia, nombre, codigo, email in asistencias:
                # Calculate worked hours
                horas_trabajadas = 0
                if asistencia.hora_entrada and asistencia.hora_salida:
                    entrada = datetime.combine(date.today(), asistencia.hora_entrada)
                    salida = datetime.combine(date.today(), asistencia.hora_salida)
                    diferencia = salida - entrada
                    horas_trabajadas = diferencia.total_seconds() / 3600
                
                # Get expected hours
                horario = db.query(Horario).filter(
                    Horario.user_id == asistencia.user_id,
                    Horario.activo == True
                ).first()
                
                horas_requeridas = horario.horas_requeridas / 60 if horario else 8
                
                # Determine status
                estado_texto = self._obtener_estado_texto(asistencia.estado)
                
                datos.append({
                    "codigo": codigo,
                    "nombre": nombre,
                    "email": email,
                    "fecha": asistencia.fecha,
                    "hora_entrada": asistencia.hora_entrada.strftime("%H:%M") if asistencia.hora_entrada else "-",
                    "hora_salida": asistencia.hora_salida.strftime("%H:%M") if asistencia.hora_salida else "-",
                    "horas_trabajadas": round(horas_trabajadas, 2),
                    "horas_requeridas": round(horas_requeridas, 2),
                    "diferencia_horas": round(horas_trabajadas - horas_requeridas, 2),
                    "estado": estado_texto,
                    "tardanza": asistencia.tardanza,
                    "minutos_tardanza": asistencia.minutos_tardanza or 0,
                    "justificacion": asistencia.observaciones or "-"
                })
            
            return datos
            
        except Exception as e:
            logger.error(f"Error calculating attendance data: {str(e)}")
            return []
    
    def _obtener_estado_texto(self, estado: EstadoAsistencia) -> str:
        """Convert status enum to readable text"""
        estados = {
            EstadoAsistencia.PRESENTE: "Presente",
            EstadoAsistencia.AUSENTE: "Ausente",
            EstadoAsistencia.TARDE: "Tarde",
            EstadoAsistencia.JUSTIFICADO: "Justificado",
            EstadoAsistencia.PERMISO: "Permiso"
        }
        return estados.get(estado, "Desconocido")
    
    def _calcular_estadisticas(self, datos: List[Dict[str, Any]]) -> Dict[str, Any]:
        """Calculate statistics from attendance data"""
        total_registros = len(datos)
        
        if total_registros == 0:
            return {
                "total_registros": 0,
                "total_presentes": 0,
                "total_ausentes": 0,
                "total_tardanzas": 0,
                "total_justificados": 0,
                "promedio_horas": 0,
                "total_minutos_tardanza": 0
            }
        
        presentes = sum(1 for d in datos if d["estado"] == "Presente")
        ausentes = sum(1 for d in datos if d["estado"] == "Ausente")
        tardanzas = sum(1 for d in datos if d["tardanza"])
        justificados = sum(1 for d in datos if d["estado"] == "Justificado")
        
        total_horas = sum(d["horas_trabajadas"] for d in datos)
        promedio_horas = total_horas / total_registros if total_registros > 0 else 0
        
        total_minutos_tardanza = sum(d["minutos_tardanza"] for d in datos)
        
        return {
            "total_registros": total_registros,
            "total_presentes": presentes,
            "total_ausentes": ausentes,
            "total_tardanzas": tardanzas,
            "total_justificados": justificados,
            "promedio_horas": round(promedio_horas, 2),
            "total_minutos_tardanza": total_minutos_tardanza
        }
    
    # ========================
    # EXCEL GENERATION
    # ========================
    
    def _generar_excel(
        self,
        datos: List[Dict[str, Any]],
        estadisticas: Dict[str, Any],
        titulo: str,
        periodo: str
    ) -> str:
        """
        Generate Excel report
        Requerimiento #14
        """
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Reporte Asistencia"
            
            # Styles
            title_font = Font(bold=True, size=14)
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Title
            ws.merge_cells('A1:M1')
            title_cell = ws['A1']
            title_cell.value = titulo
            title_cell.font = title_font
            title_cell.alignment = Alignment(horizontal='center')
            
            # Period
            ws.merge_cells('A2:M2')
            period_cell = ws['A2']
            period_cell.value = f"Período: {periodo}"
            period_cell.alignment = Alignment(horizontal='center')
            
            # Empty row
            ws.append([])
            
            # Headers
            headers = [
                "Código",
                "Nombre",
                "Fecha",
                "Entrada",
                "Salida",
                "Horas Trabajadas",
                "Horas Requeridas",
                "Diferencia",
                "Estado",
                "Tardanza",
                "Min. Tarde",
                "Justificación"
            ]
            ws.append(headers)
            
            # Style headers
            header_row = ws[4]
            for cell in header_row:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')
                cell.border = border
            
            # Data rows
            for dato in datos:
                row = [
                    dato["codigo"],
                    dato["nombre"],
                    dato["fecha"].strftime("%d/%m/%Y"),
                    dato["hora_entrada"],
                    dato["hora_salida"],
                    dato["horas_trabajadas"],
                    dato["horas_requeridas"],
                    dato["diferencia_horas"],
                    dato["estado"],
                    "Sí" if dato["tardanza"] else "No",
                    dato["minutos_tardanza"],
                    dato["justificacion"]
                ]
                ws.append(row)
                
                # Apply border to all cells in row
                for cell in ws[ws.max_row]:
                    cell.border = border
            
            # Add statistics
            ws.append([])
            ws.append(["ESTADÍSTICAS"])
            stats_start_row = ws.max_row
            
            ws.append(["Total de registros:", estadisticas["total_registros"]])
            ws.append(["Total presentes:", estadisticas["total_presentes"]])
            ws.append(["Total ausentes:", estadisticas["total_ausentes"]])
            ws.append(["Total tardanzas:", estadisticas["total_tardanzas"]])
            ws.append(["Total justificados:", estadisticas["total_justificados"]])
            ws.append(["Promedio de horas trabajadas:", estadisticas["promedio_horas"]])
            ws.append(["Total minutos de tardanza:", estadisticas["total_minutos_tardanza"]])
            
            # Bold statistics
            for row in range(stats_start_row, ws.max_row + 1):
                ws.cell(row, 1).font = Font(bold=True)
            
            # Adjust column widths
            for col in range(1, 13):
                ws.column_dimensions[get_column_letter(col)].width = 15
            
            # Save file
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"reporte_asistencia_{timestamp}.xlsx"
            filepath = self.reports_dir / filename
            
            wb.save(filepath)
            logger.info(f"Excel report generated: {filepath}")
            
            return str(filepath)
            
        except Exception as e:
            logger.error(f"Error generating Excel report: {str(e)}")
            raise
    
    # ========================
    # PDF GENERATION
    # ========================
    
    def _generar_pdf(
        self,
        datos: List[Dict[str, Any]],
        estadisticas: Dict[str, Any],
        titulo: str,
        periodo: str
    ) -> str:
        """
        Generate PDF report
        Requerimiento #14
        """
        try:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"reporte_asistencia_{timestamp}.pdf"
            filepath = self.reports_dir / filename
            
            doc = SimpleDocTemplate(str(filepath), pagesize=A4)
            elements = []
            styles = getSampleStyleSheet()
            
            # Custom styles
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=16,
                textColor=colors.HexColor('#366092'),
                spaceAfter=12,
                alignment=TA_CENTER
            )
            
            subtitle_style = ParagraphStyle(
                'CustomSubtitle',
                parent=styles['Normal'],
                fontSize=12,
                spaceAfter=12,
                alignment=TA_CENTER
            )
            
            # Title
            elements.append(Paragraph(titulo, title_style))
            elements.append(Paragraph(f"Período: {periodo}", subtitle_style))
            elements.append(Spacer(1, 0.2 * inch))
            
            # Statistics table
            stats_data = [
                ["ESTADÍSTICAS", ""],
                ["Total de registros:", str(estadisticas["total_registros"])],
                ["Total presentes:", str(estadisticas["total_presentes"])],
                ["Total ausentes:", str(estadisticas["total_ausentes"])],
                ["Total tardanzas:", str(estadisticas["total_tardanzas"])],
                ["Promedio de horas:", f"{estadisticas['promedio_horas']} hrs"],
                ["Total minutos tardanza:", str(estadisticas["total_minutos_tardanza"])]
            ]
            
            stats_table = Table(stats_data, colWidths=[3 * inch, 2 * inch])
            stats_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            elements.append(stats_table)
            elements.append(Spacer(1, 0.3 * inch))
            
            # Data table
            if datos:
                # Headers
                table_data = [[
                    "Código",
                    "Nombre",
                    "Fecha",
                    "Entrada",
                    "Salida",
                    "Hrs",
                    "Estado"
                ]]
                
                # Data rows (simplified for PDF)
                for dato in datos:
                    row = [
                        dato["codigo"],
                        dato["nombre"][:20],  # Truncate long names
                        dato["fecha"].strftime("%d/%m/%Y"),
                        dato["hora_entrada"],
                        dato["hora_salida"],
                        str(dato["horas_trabajadas"]),
                        dato["estado"][:10]
                    ]
                    table_data.append(row)
                
                # Create table
                data_table = Table(table_data, colWidths=[0.8*inch, 1.5*inch, 0.9*inch, 0.7*inch, 0.7*inch, 0.6*inch, 1*inch])
                data_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 10),
                    ('FONTSIZE', (0, 1), (-1, -1), 8),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.white),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black),
                    ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey])
                ]))
                
                elements.append(data_table)
            
            # Build PDF
            doc.build(elements)
            logger.info(f"PDF report generated: {filepath}")
            
            return str(filepath)
            
        except Exception as e:
            logger.error(f"Error generating PDF report: {str(e)}")
            raise
    
    # ========================
    # PUBLIC REPORT METHODS
    # ========================
    
    async def generar_reporte_diario(
        self,
        db: Session,
        fecha: date,
        user_id: Optional[int] = None,
        formato: str = "both",  # "pdf", "excel", "both"
        enviar_email: bool = False
    ) -> Dict[str, Any]:
        """
        Generate daily attendance report
        Requerimiento #11
        """
        try:
            titulo = "REPORTE DIARIO DE ASISTENCIA"
            periodo = fecha.strftime("%d/%m/%Y")
            
            # Get data
            datos = self._calcular_datos_asistencia(db, user_id, fecha, fecha)
            estadisticas = self._calcular_estadisticas(datos)
            
            # Generate files
            archivos = {}
            
            if formato in ["excel", "both"]:
                excel_path = self._generar_excel(datos, estadisticas, titulo, periodo)
                archivos["excel"] = excel_path
            
            if formato in ["pdf", "both"]:
                pdf_path = self._generar_pdf(datos, estadisticas, titulo, periodo)
                archivos["pdf"] = pdf_path
            
            # Send email if requested
            if enviar_email:
                try:
                    # Get user email(s)
                    if user_id:
                        user = db.query(User).filter(User.id == user_id).first()
                        emails = [user.email] if user and user.email else []
                    else:
                        # Send to all admins - usando join con Role
                        admins = db.query(User).join(Role).filter(Role.es_admin == True).all()
                        emails = [admin.email for admin in admins if admin.email]
                    
                    if emails:
                        await email_service.send_report_email(
                            to_email=emails,
                            report_type="diario",
                            period=periodo,
                            attachments=self._preparar_adjuntos(archivos)
                        )
                        logger.info(f"Report email sent to {len(emails)} recipients")
                except Exception as email_error:
                    logger.error(f"Error sending report email: {str(email_error)}")
            
            return {
                "success": True,
                "tipo": "diario",
                "periodo": periodo,
                "archivos": archivos,
                "estadisticas": estadisticas,
                "total_registros": len(datos)
            }
            
        except Exception as e:
            logger.error(f"Error generating daily report: {str(e)}")
            return {"success": False, "error": str(e)}
    
    async def generar_reporte_semanal(
        self,
        db: Session,
        fecha_inicio: date,
        user_id: Optional[int] = None,
        formato: str = "both",
        enviar_email: bool = False
    ) -> Dict[str, Any]:
        """
        Generate weekly attendance report
        Requerimiento #12
        """
        try:
            fecha_fin = fecha_inicio + timedelta(days=6)
            titulo = "REPORTE SEMANAL DE ASISTENCIA"
            periodo = f"{fecha_inicio.strftime('%d/%m/%Y')} - {fecha_fin.strftime('%d/%m/%Y')}"
            
            # Get data
            datos = self._calcular_datos_asistencia(db, user_id, fecha_inicio, fecha_fin)
            estadisticas = self._calcular_estadisticas(datos)
            
            # Generate files
            archivos = {}
            
            if formato in ["excel", "both"]:
                excel_path = self._generar_excel(datos, estadisticas, titulo, periodo)
                archivos["excel"] = excel_path
            
            if formato in ["pdf", "both"]:
                pdf_path = self._generar_pdf(datos, estadisticas, titulo, periodo)
                archivos["pdf"] = pdf_path
            
            # Send email if requested
            if enviar_email:
                try:
                    if user_id:
                        user = db.query(User).filter(User.id == user_id).first()
                        emails = [user.email] if user and user.email else []
                    else:
                        admins = db.query(User).join(Role).filter(Role.es_admin == True).all()
                        emails = [admin.email for admin in admins if admin.email]
                    
                    if emails:
                        await email_service.send_report_email(
                            to_email=emails,
                            report_type="semanal",
                            period=periodo,
                            attachments=self._preparar_adjuntos(archivos)
                        )
                        logger.info(f"Weekly report email sent to {len(emails)} recipients")
                except Exception as email_error:
                    logger.error(f"Error sending weekly report email: {str(email_error)}")
            
            return {
                "success": True,
                "tipo": "semanal",
                "periodo": periodo,
                "archivos": archivos,
                "estadisticas": estadisticas,
                "total_registros": len(datos)
            }
            
        except Exception as e:
            logger.error(f"Error generating weekly report: {str(e)}")
            return {"success": False, "error": str(e)}
    
    async def generar_reporte_mensual(
        self,
        db: Session,
        anio: int,
        mes: int,
        user_id: Optional[int] = None,
        formato: str = "both",
        enviar_email: bool = False
    ) -> Dict[str, Any]:
        """
        Generate monthly attendance report
        Requerimiento #13
        """
        try:
            fecha_inicio = date(anio, mes, 1)
            
            # Calculate last day of month
            if mes == 12:
                fecha_fin = date(anio + 1, 1, 1) - timedelta(days=1)
            else:
                fecha_fin = date(anio, mes + 1, 1) - timedelta(days=1)
            
            titulo = "REPORTE MENSUAL DE ASISTENCIA"
            periodo = fecha_inicio.strftime("%B %Y")
            
            # Get data
            datos = self._calcular_datos_asistencia(db, user_id, fecha_inicio, fecha_fin)
            estadisticas = self._calcular_estadisticas(datos)
            
            # Generate files
            archivos = {}
            
            if formato in ["excel", "both"]:
                excel_path = self._generar_excel(datos, estadisticas, titulo, periodo)
                archivos["excel"] = excel_path
            
            if formato in ["pdf", "both"]:
                pdf_path = self._generar_pdf(datos, estadisticas, titulo, periodo)
                archivos["pdf"] = pdf_path
            
            # Send email if requested
            if enviar_email:
                try:
                    if user_id:
                        user = db.query(User).filter(User.id == user_id).first()
                        emails = [user.email] if user and user.email else []
                    else:
                        admins = db.query(User).join(Role).filter(Role.es_admin == True).all()
                        emails = [admin.email for admin in admins if admin.email]
                    
                    if emails:
                        await email_service.send_report_email(
                            to_email=emails,
                            report_type="mensual",
                            period=periodo,
                            attachments=self._preparar_adjuntos(archivos)
                        )
                        logger.info(f"Monthly report email sent to {len(emails)} recipients")
                except Exception as email_error:
                    logger.error(f"Error sending monthly report email: {str(email_error)}")
            
            return {
                "success": True,
                "tipo": "mensual",
                "periodo": periodo,
                "archivos": archivos,
                "estadisticas": estadisticas,
                "total_registros": len(datos)
            }
            
        except Exception as e:
            logger.error(f"Error generating monthly report: {str(e)}")
            return {"success": False, "error": str(e)}
    
    async def generar_reporte_tardanzas(
        self,
        db: Session,
        fecha_inicio: date,
        fecha_fin: date,
        user_id: Optional[int] = None,
        formato: str = "both",
        enviar_email: bool = False
    ) -> Dict[str, Any]:
        """
        Generate consolidated late arrivals report
        Requerimiento #14
        """
        try:
            titulo = "REPORTE CONSOLIDADO DE TARDANZAS"
            periodo = f"{fecha_inicio.strftime('%d/%m/%Y')} - {fecha_fin.strftime('%d/%m/%Y')}"
            
            # Get all data and filter late arrivals
            todos_datos = self._calcular_datos_asistencia(db, user_id, fecha_inicio, fecha_fin)
            datos = [d for d in todos_datos if d["tardanza"]]
            
            estadisticas = self._calcular_estadisticas(datos)
            
            # Generate files
            archivos = {}
            
            if formato in ["excel", "both"]:
                excel_path = self._generar_excel(datos, estadisticas, titulo, periodo)
                archivos["excel"] = excel_path
            
            if formato in ["pdf", "both"]:
                pdf_path = self._generar_pdf(datos, estadisticas, titulo, periodo)
                archivos["pdf"] = pdf_path
            
            # Send email if requested
            if enviar_email:
                try:
                    if user_id:
                        user = db.query(User).filter(User.id == user_id).first()
                        emails = [user.email] if user and user.email else []
                    else:
                        admins = db.query(User).join(Role).filter(Role.es_admin == True).all()
                        emails = [admin.email for admin in admins if admin.email]
                    
                    if emails:
                        await email_service.send_report_email(
                            to_email=emails,
                            report_type="tardanzas",
                            period=periodo,
                            attachments=self._preparar_adjuntos(archivos)
                        )
                        logger.info(f"Late arrivals report email sent to {len(emails)} recipients")
                except Exception as email_error:
                    logger.error(f"Error sending late arrivals report email: {str(email_error)}")
            
            return {
                "success": True,
                "tipo": "tardanzas",
                "periodo": periodo,
                "archivos": archivos,
                "estadisticas": estadisticas,
                "total_registros": len(datos)
            }
            
        except Exception as e:
            logger.error(f"Error generating late arrivals report: {str(e)}")
            return {"success": False, "error": str(e)}
    
    async def generar_reporte_inasistencias(
        self,
        db: Session,
        fecha_inicio: date,
        fecha_fin: date,
        user_id: Optional[int] = None,
        formato: str = "both",
        enviar_email: bool = False
    ) -> Dict[str, Any]:
        """
        Generate consolidated absences report
        Requerimiento #15
        """
        try:
            titulo = "REPORTE CONSOLIDADO DE INASISTENCIAS"
            periodo = f"{fecha_inicio.strftime('%d/%m/%Y')} - {fecha_fin.strftime('%d/%m/%Y')}"
            
            # Get all data and filter absences
            todos_datos = self._calcular_datos_asistencia(db, user_id, fecha_inicio, fecha_fin)
            datos = [d for d in todos_datos if d["estado"] == "Ausente"]
            
            estadisticas = self._calcular_estadisticas(datos)
            
            # Generate files
            archivos = {}
            
            if formato in ["excel", "both"]:
                excel_path = self._generar_excel(datos, estadisticas, titulo, periodo)
                archivos["excel"] = excel_path
            
            if formato in ["pdf", "both"]:
                pdf_path = self._generar_pdf(datos, estadisticas, titulo, periodo)
                archivos["pdf"] = pdf_path
            
            # Send email if requested
            if enviar_email:
                try:
                    if user_id:
                        user = db.query(User).filter(User.id == user_id).first()
                        emails = [user.email] if user and user.email else []
                    else:
                        admins = db.query(User).join(Role).filter(Role.es_admin == True).all()
                        emails = [admin.email for admin in admins if admin.email]
                    
                    if emails:
                        await email_service.send_report_email(
                            to_email=emails,
                            report_type="inasistencias",
                            period=periodo,
                            attachments=self._preparar_adjuntos(archivos)
                        )
                        logger.info(f"Absences report email sent to {len(emails)} recipients")
                except Exception as email_error:
                    logger.error(f"Error sending absences report email: {str(email_error)}")
            
            return {
                "success": True,
                "tipo": "inasistencias",
                "periodo": periodo,
                "archivos": archivos,
                "estadisticas": estadisticas,
                "total_registros": len(datos)
            }
            
        except Exception as e:
            logger.error(f"Error generating absences report: {str(e)}")
            return {"success": False, "error": str(e)}
    
    async def enviar_reporte_por_correo(
        self,
        reporte_resultado: Dict[str, Any],
        destinatarios: List[str]
    ) -> bool:
        """
        Send generated report by email
        """
        try:
            if not reporte_resultado.get("success"):
                return False
            
            archivos = reporte_resultado.get("archivos", {})
            attachments = []
            
            # Prepare attachments
            for tipo, filepath in archivos.items():
                if Path(filepath).exists():
                    filename = Path(filepath).name
                    attachments.append({
                        "filename": filename,
                        "path": filepath
                    })
            
            # Send email
            return await email_service.send_report_email(
                to_email=destinatarios,
                report_type=reporte_resultado["tipo"],
                period=reporte_resultado["periodo"],
                attachments=attachments
            )
            
        except Exception as e:
            logger.error(f"Error sending report by email: {str(e)}")
            return False


# Singleton instance
reportes_service = ReportesService()
